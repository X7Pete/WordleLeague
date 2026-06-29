from openpyxl import load_workbook
from config import WORKBOOK_NAME


def update_streaks():
    wb = load_workbook(WORKBOOK_NAME)
    results_ws = wb["Results"]
    streaks_ws = wb["Streaks"]

    # Läs rubriker
    headers = [cell.value for cell in results_ws[1]]
    player_names = headers[2:]

    # Samla alla rader sorterade på datum
    all_rows = list(results_ws.iter_rows(min_row=2, values_only=True))
    all_rows.sort(key=lambda x: x[0])  # Sortera på datum

    # Bygg historik per spelare: lista av (date, score, won)
    history = {name: [] for name in player_names}

    for row in all_rows:
        date = row[0]
        scores = {}
        for i, name in enumerate(player_names):
            score = row[2 + i]
            if score is not None:
                scores[name] = score

        if not scores:
            continue

        best_score = min(scores.values())

        for name, score in scores.items():
            won = score == best_score
            history[name].append((date, score, won))

    # Beräkna streaks per spelare
    results = {}

    for name in player_names:
        games = history[name]

        if not games:
            results[name] = {
                "current_streak": 0,
                "best_streak": 0,
                "current_win_streak": 0,
                "best_win_streak": 0,
            }
            continue

        # --- Current streak (spelade dagar i rad från slutet) ---
        current_streak = 0
        for _, score, _ in reversed(games):
            if score is not None:
                current_streak += 1
            else:
                break

        # --- Best streak ---
        best_streak = 0
        running = 0
        for _, score, _ in games:
            if score is not None:
                running += 1
                best_streak = max(best_streak, running)
            else:
                running = 0

        # --- Current win streak ---
        current_win_streak = 0
        for _, _, won in reversed(games):
            if won:
                current_win_streak += 1
            else:
                break

        # --- Best win streak ---
        best_win_streak = 0
        running = 0
        for _, _, won in games:
            if won:
                running += 1
                best_win_streak = max(best_win_streak, running)
            else:
                running = 0

        results[name] = {
            "current_streak": current_streak,
            "best_streak": best_streak,
            "current_win_streak": current_win_streak,
            "best_win_streak": best_win_streak,
        }

    # Sortera på current streak, sen best streak
    sorted_results = sorted(
        results.items(),
        key=lambda x: (-x[1]["current_streak"], -x[1]["best_streak"])
    )

    # Skriv till bladet
    streaks_ws.delete_rows(1, streaks_ws.max_row)
    streaks_ws.append(["Name", "Current Streak", "Best Streak", "Current Win Streak", "Best Win Streak"])

    for name, s in sorted_results:
        streaks_ws.append([
            name,
            s["current_streak"],
            s["best_streak"],
            s["current_win_streak"],
            s["best_win_streak"],
        ])

    wb.save(WORKBOOK_NAME)
    print("✅ Streaks uppdaterad!")