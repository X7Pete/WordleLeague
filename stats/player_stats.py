from openpyxl import load_workbook
from config import WORKBOOK_NAME


def update_player_stats():
    wb = load_workbook(WORKBOOK_NAME)
    results_ws = wb["Results"]
    stats_ws = wb["Player Stats"]

    # Läs rubriker
    headers = [cell.value for cell in results_ws[1]]
    player_names = headers[2:]

    # Samla historik per spelare
    history = {name: [] for name in player_names}

    for row in results_ws.iter_rows(min_row=2, values_only=True):
        date = row[0]
        wordle = row[1]

        for i, name in enumerate(player_names):
            score = row[2 + i]
            if score is not None:
                history[name].append((date, wordle, score))

    # Rensa bladet
    stats_ws.delete_rows(1, stats_ws.max_row)

    # Skriv en sektion per spelare
    for name in player_names:
        games = history[name]

        # Spelarnamn som rubrik
        stats_ws.append([name])
        stats_ws.append(["Date", "Wordle", "Score"])

        for date, wordle, score in games:
            stats_ws.append([date, wordle, score])

        # Beräkna sammanfattning
        scores = [score for _, _, score in games]
        if scores:
            total_games = len(scores)
            average = round(sum(scores) / total_games, 2)
            best = min(scores)
            worst = max(scores)

            # Räkna wins
            wins = 0
            for row in results_ws.iter_rows(min_row=2, values_only=True):
                day_scores = {}
                for i, n in enumerate(player_names):
                    s = row[2 + i]
                    if s is not None:
                        day_scores[n] = s
                if day_scores and name in day_scores:
                    if day_scores[name] == min(day_scores.values()):
                        wins += 1

            stats_ws.append([])  # Tom rad
            stats_ws.append([
                f"Games: {total_games}",
                f"Wins: {wins}",
                f"Average: {average}",
                f"Best: {best}",
                f"Worst: {worst}"
            ])

        stats_ws.append([])  # Tom rad mellan spelare
        stats_ws.append([])

    wb.save(WORKBOOK_NAME)
    print("✅ Player Stats uppdaterad!")