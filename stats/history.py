from openpyxl import load_workbook
from config import WORKBOOK_NAME


def update_history():
    wb = load_workbook(WORKBOOK_NAME)
    results_ws = wb["Results"]
    history_ws = wb["History"]

    # Läs rubriker
    headers = [cell.value for cell in results_ws[1]]
    player_names = headers[2:]

    # Samla alla rader
    all_rows = list(results_ws.iter_rows(min_row=2, values_only=True))
    all_rows.sort(key=lambda x: (x[0], x[1]), reverse=True)
    # Rensa bladet
    history_ws.delete_rows(1, history_ws.max_row)
    history_ws.append(["Date", "Wordle", "1st", "2nd", "3rd", "4th", "5th", "6th"])

    for row in all_rows:
        date = row[0]
        wordle = row[1]

        scores = {}
        for i, name in enumerate(player_names):
            score = row[2 + i]
            if score is not None:
                scores[name] = score

        if not scores:
            continue

        # Sortera spelare efter score
        sorted_players = sorted(scores.items(), key=lambda x: x[1])

        # Bygg rad med placeringar
        history_row = [date, wordle]
        for name, score in sorted_players:
            history_row.append(f"{name} ({score})")

        # Fyll ut tomma platser om färre än 6 spelade
        while len(history_row) < 8:
            history_row.append("")

        history_ws.append(history_row)

    wb.save(WORKBOOK_NAME)
    print("✅ History uppdaterad!")