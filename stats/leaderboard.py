from openpyxl import load_workbook
from config import WORKBOOK_NAME


def update_leaderboard():
    wb = load_workbook(WORKBOOK_NAME)
    results_ws = wb["Results"]
    leaderboard_ws = wb["Leaderboard"]

    # Läs rubriker från Results-bladet
    headers = [cell.value for cell in results_ws[1]]
    player_names = headers[2:]  # Hoppa över Date och Wordle

    # Samla statistik per spelare
    stats = {
        name: {
            "points": 0,
            "total": 0,
            "games": 0,
            "wins": 0,
            "podiums": 0,
            "best": None,
            "worst": None,
        }
        for name in player_names
    }

    # Gå igenom varje rad (dag)
    for row in results_ws.iter_rows(min_row=2, values_only=True):
        scores = {}

        for i, name in enumerate(player_names):
            score = row[2 + i]
            if score is not None:
                scores[name] = score

        if not scores:
            continue

        # --- Beräkna Points med tie-hantering ---
        sorted_players = sorted(scores.items(), key=lambda x: x[1])
        point_map = [10, 7, 5, 3, 2, 1]
        prev_score = None
        rank = 0

        for i, (name, score) in enumerate(sorted_players):
            if score != prev_score:
                rank = i
                prev_score = score
            if rank < len(point_map):
                stats[name]["points"] += point_map[rank]

        # --- Wins ---
        best_score = min(scores.values())
        for name, score in scores.items():
            if score == best_score:
                stats[name]["wins"] += 1

        # --- Podiums (topp 3) ---
        unique_scores = sorted(set(scores.values()))
        top3_scores = set(unique_scores[:3])
        for name, score in scores.items():
            if score in top3_scores:
                stats[name]["podiums"] += 1

        # --- Average, Best, Worst ---
        for name, score in scores.items():
            stats[name]["total"] += score
            stats[name]["games"] += 1

            if stats[name]["best"] is None or score < stats[name]["best"]:
                stats[name]["best"] = score

            if stats[name]["worst"] is None or score > stats[name]["worst"]:
                stats[name]["worst"] = score

    # Sortera: Points fallande, sen Average stigande
    sorted_stats = sorted(
        stats.items(),
        key=lambda x: (
            -x[1]["points"],
            x[1]["total"] / x[1]["games"] if x[1]["games"] > 0 else 999
        )
    )

    # Skriv leaderboard
    leaderboard_ws.delete_rows(1, leaderboard_ws.max_row)
    leaderboard_ws.append(["Rank", "Name", "Points", "Games", "Wins", "Podiums", "Best", "Worst", "Average"])

    for rank, (name, s) in enumerate(sorted_stats, start=1):
        avg = round(s["total"] / s["games"], 2) if s["games"] > 0 else "-"
        leaderboard_ws.append([
            rank,
            name,
            s["points"],
            s["games"],
            s["wins"],
            s["podiums"],
            s["best"],
            s["worst"],
            avg
        ])

    wb.save(WORKBOOK_NAME)
    print("✅ Leaderboard uppdaterad!")