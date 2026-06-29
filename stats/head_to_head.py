from openpyxl import load_workbook
from config import WORKBOOK_NAME


def update_head_to_head():
    wb = load_workbook(WORKBOOK_NAME)
    results_ws = wb["Results"]
    h2h_ws = wb["Head-to-Head"]

    # Läs rubriker
    headers = [cell.value for cell in results_ws[1]]
    player_names = headers[2:]

    # Bygg H2H-matris: h2h[a][b] = [wins, losses, draws]
    h2h = {a: {b: [0, 0, 0] for b in player_names if b != a} for a in player_names}

    for row in results_ws.iter_rows(min_row=2, values_only=True):
        scores = {}
        for i, name in enumerate(player_names):
            score = row[2 + i]
            if score is not None:
                scores[name] = score

        if not scores:
            continue

        # Jämför varje par
        players_today = list(scores.keys())
        for i in range(len(players_today)):
            for j in range(len(players_today)):
                if i == j:
                    continue
                a = players_today[i]
                b = players_today[j]
                if scores[a] < scores[b]:
                    h2h[a][b][0] += 1
                elif scores[a] > scores[b]:
                    h2h[a][b][1] += 1
                else:
                    h2h[a][b][2] += 1  # Oavgjort

    # Rensa bladet
    h2h_ws.delete_rows(1, h2h_ws.max_row)

    # Rubrikrad
    h2h_ws.append([""] + player_names)

    # Skriv en rad per spelare
    for a in player_names:
        row = [a]
        for b in player_names:
            if a == b:
                row.append("-")
            else:
                wins, losses, draws = h2h[a][b]
                row.append(f"{wins}-{losses}-{draws}")
        h2h_ws.append(row)

    wb.save(WORKBOOK_NAME)
    print("✅ Head-to-Head uppdaterad!")