from openpyxl import load_workbook
from config import WORKBOOK_NAME

def clear_results(players):
    wb = load_workbook(WORKBOOK_NAME)
    ws = wb["Results"]
    
    # Radera allt utom rubrikraden
    ws.delete_rows(2, ws.max_row)
    
    wb.save(WORKBOOK_NAME)

def add_result(players, date, wordle, results):

    workbook = load_workbook("WordleLeague.xlsx")

    sheet = workbook["Results"]

    row = [
        date,
        wordle
    ]

    for player in players:
        row.append(results.get(player, ""))


    sheet.append(row)

    workbook.save("WordleLeague.xlsx")