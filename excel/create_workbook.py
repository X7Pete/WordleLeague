from openpyxl import Workbook
from pathlib import Path
from config import WORKBOOK_NAME

def create_workbook(players):
    filename = WORKBOOK_NAME

    if filename.exists():
        print("Filen WordleLeague.xlsx är redan skapad!")
        return

    workbook = Workbook()

    # Ta bort standardbladet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Skapa blad
    workbook.create_sheet("Dashboard")
    workbook.create_sheet("Results")
    workbook.create_sheet("Leaderboard")
    workbook.create_sheet("Player Stats")
    workbook.create_sheet("Streaks")
    workbook.create_sheet("Head-to-Head")
    workbook.create_sheet("History")
    workbook.create_sheet("Settings")

    # Rubriker i Results
    results_sheet = workbook["Results"]
    results_sheet.append(["Date", "Wordle"] + players)

    # Rubriker i Leaderboard
    leaderboard_sheet = workbook["Leaderboard"]
    leaderboard_sheet.append(["Rank", "Name", "Points", "Games", "Wins", "Average"])

    # Spara EN gång, när allt är klart
    workbook.save(WORKBOOK_NAME)
    print("✅ WordleLeague.xlsx skapad!")