from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
from config import WORKBOOK_NAME


def update_dashboard():
    wb = load_workbook(WORKBOOK_NAME)
    results_ws = wb["Results"]
    leaderboard_ws = wb["Leaderboard"]
    dash_ws = wb["Dashboard"]

    # Rensa bladet
    dash_ws.delete_rows(1, dash_ws.max_row)

    # Läs spelarnamn
    headers = [cell.value for cell in results_ws[1]]
    player_names = headers[2:]

    # Läs all resultatdata
    all_rows = list(results_ws.iter_rows(min_row=2, values_only=True))

    # --- Titel ---
    dash_ws.append(["🏆 WordleLeague Dashboard"])
    dash_ws["A1"].font = Font(bold=True, size=16)
    dash_ws.append([f"Senast uppdaterad: {date.today()}"])
    dash_ws.append([])

    # --- Leaderboard-sammanfattning ---
    dash_ws.append(["LEADERBOARD"])
    dash_ws[f"A{dash_ws.max_row}"].font = Font(bold=True, size=12)
    dash_ws.append(["Rank", "Name", "Points", "Games", "Wins", "Average"])

    for row in leaderboard_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            dash_ws.append(list(row[:6]))  # Rank, Name, Points, Games, Wins, Average

    dash_ws.append([])

    # --- Roliga fakta ---
    dash_ws.append(["ROLIGA FAKTA"])
    dash_ws[f"A{dash_ws.max_row}"].font = Font(bold=True, size=12)

    # Bästa och sämsta resultat någonsin
    best_score = None
    worst_score = None
    best_players = []
    worst_players = []

    for row in all_rows:
        for i, name in enumerate(player_names):
            score = row[2 + i]
            if score is None:
                continue
            if best_score is None or score < best_score:
                best_score = score
                best_players = [name]          # Nollställ listan
            elif score == best_score:
                if name not in best_players:   # 👈 Undvik dubletter
                    best_players.append(name)
            if worst_score is None or score > worst_score:
                worst_score = score
                worst_players = [name]         # Nollställ listan
            elif score == worst_score:
                if name not in worst_players:  # 👈 Undvik dubletter
                    worst_players.append(name)

    # Mest wins från leaderboard
    win_rows = list(leaderboard_ws.iter_rows(min_row=2, values_only=True))
    if win_rows:
        max_wins = max(r[4] for r in win_rows if r[4] is not None)
        top_winners = [r[1] for r in win_rows if r[4] == max_wins]
    else:
        max_wins = 0
        top_winners = []

    dash_ws.append([f"🎯 Bästa resultat någonsin:", f"{' & '.join(best_players)} ({best_score})"])
    dash_ws.append([f"📉 Sämsta resultat någonsin:", f"{' & '.join(worst_players)} ({worst_score})"])
    dash_ws.append([f"🏆 Mest wins:", f"{' & '.join(top_winners)} ({max_wins} wins)"])

    # Senaste dagens resultat
    if all_rows:
        last_row = all_rows[-1]
        last_date = last_row[0]
        last_scores = {}
        for i, name in enumerate(player_names):
            score = last_row[2 + i]
            if score is not None:
                last_scores[name] = score

        if last_scores:
            best_today = min(last_scores.values())
            heroes = [n for n, s in last_scores.items() if s == best_today]
            dash_ws.append([f"⭐ Senaste dagens hjälte:", f"{' & '.join(heroes)} ({best_today})"])

    wb.save(WORKBOOK_NAME)
    print("✅ Dashboard uppdaterad!")